from datetime import datetime, date
import io

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..database import get_db
from ..deps import require_branch_access, require_perm, has_perm
from ..models import Business, DisbursementBatch, DisbursementItem, Program
from ..utils import paginate, parse_page

router = APIRouter(tags=["disbursements"])


def batch_ser(b: DisbursementBatch, db: Session, include_items: bool = True) -> dict:
    data = {
        "id": b.id,
        "branch_id": b.branch_id,
        "program_id": b.program_id,
        "program_name": b.program.name if b.program else None,
        "title": b.title,
        "payment_date": str(b.payment_date) if b.payment_date else None,
        "currency": b.currency,
        "base_amount": float(b.base_amount or 0.0),
        "total_amount": float(b.total_amount or 0.0),
        "status": b.status or "draft",
        "notes": b.notes or "",
        "created_by": {
            "id": b.created_by.id,
            "name": f"{b.created_by.first_name} {b.created_by.last_name}".strip() or b.created_by.email,
            "email": b.created_by.email,
        } if b.created_by else None,
        "confirmed_by": {
            "id": b.confirmed_by.id,
            "name": f"{b.confirmed_by.first_name} {b.confirmed_by.last_name}".strip() or b.confirmed_by.email,
            "email": b.confirmed_by.email,
        } if b.confirmed_by else None,
        "confirmed_at": str(b.confirmed_at) if b.confirmed_at else None,
        "created_at": str(b.created_at) if b.created_at else None,
        "updated_at": str(b.updated_at) if b.updated_at else None,
        "items_count": len(b.items) if b.items else 0,
    }
    if include_items:
        data["items"] = [
            {
                "id": it.id,
                "business_id": it.business_id,
                "business_name": it.business.name if it.business else "",
                "business_logo": it.business.logo if it.business else None,
                "founders": [
                    {
                        "id": f.id,
                        "name": f"{f.first_name} {f.last_name}".strip(),
                        "email": f.email,
                    }
                    for f in (it.business.founders if it.business else [])
                ],
                "percentage": float(it.percentage if it.percentage is not None else 100.0),
                "amount": float(it.amount or 0.0),
                "is_included": bool(it.is_included),
                "notes": it.notes or "",
            }
            for it in b.items
        ]
    return data


# ------------------------------------------------------------------ Summary KPIs

@router.get("/disbursements/{branch_id}/summary/")
def disbursements_summary(branch_id: int, ctx=Depends(require_perm("disbursements.view"))):
    db: Session = ctx["db"]
    batches = db.query(DisbursementBatch).filter(DisbursementBatch.branch_id == branch_id).all()
    
    total_batches = len(batches)
    draft_batches = sum(1 for b in batches if b.status == "draft")
    processed_batches = sum(1 for b in batches if b.status == "processed")
    
    totals_by_currency = {"USD": 0.0, "EUR": 0.0, "JOD": 0.0}
    pending_by_currency = {"USD": 0.0, "EUR": 0.0, "JOD": 0.0}

    for b in batches:
        curr = (b.currency or "USD").upper()
        if curr not in totals_by_currency:
            totals_by_currency[curr] = 0.0
            pending_by_currency[curr] = 0.0
        if b.status == "processed":
            totals_by_currency[curr] += float(b.total_amount or 0.0)
        else:
            pending_by_currency[curr] += float(b.total_amount or 0.0)

    return {
        "total_batches": total_batches,
        "draft_batches": draft_batches,
        "processed_batches": processed_batches,
        "totals_by_currency": totals_by_currency,
        "pending_by_currency": pending_by_currency,
    }


# ------------------------------------------------------------------ List batches

@router.get("/disbursements/{branch_id}/")
def list_disbursements(
    branch_id: int,
    request: Request,
    ctx=Depends(require_perm("disbursements.view")),
):
    db: Session = ctx["db"]
    page, size = parse_page(request)
    q = db.query(DisbursementBatch).filter(DisbursementBatch.branch_id == branch_id)

    if program_id := request.query_params.get("program_id"):
        try:
            q = q.filter(DisbursementBatch.program_id == int(program_id))
        except ValueError:
            pass

    if status := request.query_params.get("status"):
        q = q.filter(DisbursementBatch.status == status)

    if currency := request.query_params.get("currency"):
        q = q.filter(DisbursementBatch.currency == currency.upper())

    if search := request.query_params.get("search"):
        q = q.filter(DisbursementBatch.title.ilike(f"%{search}%"))

    q = q.order_by(DisbursementBatch.payment_date.desc(), DisbursementBatch.id.desc())
    return paginate(q, page, size, lambda b: batch_ser(b, db, include_items=False))


# ------------------------------------------------------------------ Create batch

class DisbursementItemIn(BaseModel):
    business_id: int
    percentage: float = 100.0
    is_included: bool = True
    notes: str = ""
    amount: float | None = None


class DisbursementBatchCreateIn(BaseModel):
    program_id: int
    title: str | None = None
    payment_date: date
    currency: str = "USD"
    base_amount: float = 0.0
    notes: str = ""
    items: list[DisbursementItemIn] | None = None


@router.post("/disbursements/{branch_id}/")
def create_disbursement_batch(
    branch_id: int,
    data: DisbursementBatchCreateIn,
    ctx=Depends(require_perm("disbursements.create")),
):
    db: Session = ctx["db"]
    user = ctx["user"]

    program = db.get(Program, data.program_id)
    if not program or program.branch_id != branch_id:
        raise HTTPException(status_code=404, detail="Programme / Cohort not found in this branch.")

    currency = data.currency.upper() if data.currency else "USD"
    if currency not in ("USD", "EUR", "JOD"):
        raise HTTPException(status_code=400, detail="Currency must be USD, EUR, or JOD.")

    title = (data.title or "").strip()
    if not title:
        month_str = data.payment_date.strftime("%B %Y")
        title = f"{program.name} — {month_str} Disbursement"

    batch = DisbursementBatch(
        branch_id=branch_id,
        program_id=data.program_id,
        title=title,
        payment_date=data.payment_date,
        currency=currency,
        base_amount=float(data.base_amount or 0.0),
        notes=data.notes or "",
        status="draft",
        created_by_id=user.id,
    )
    db.add(batch)
    db.flush()

    total_amount = 0.0

    if data.items is not None and len(data.items) > 0:
        for it in data.items:
            biz = db.get(Business, it.business_id)
            if not biz or biz.branch_id != branch_id:
                continue
            pct = float(it.percentage if it.percentage is not None else 100.0)
            item_amount = it.amount if it.amount is not None else (batch.base_amount * (pct / 100.0))
            item_amount = round(float(item_amount), 2)
            if it.is_included:
                total_amount += item_amount
            item = DisbursementItem(
                batch_id=batch.id,
                business_id=biz.id,
                percentage=pct,
                amount=item_amount,
                is_included=it.is_included,
                notes=it.notes or "",
            )
            db.add(item)
    else:
        # Default: auto-populate all businesses in this cohort / program
        businesses = db.query(Business).filter(
            Business.branch_id == branch_id,
            Business.program_id == data.program_id,
        ).order_by(Business.name).all()

        for biz in businesses:
            pct = 100.0
            item_amount = round(float(batch.base_amount * (pct / 100.0)), 2)
            total_amount += item_amount
            item = DisbursementItem(
                batch_id=batch.id,
                business_id=biz.id,
                percentage=pct,
                amount=item_amount,
                is_included=True,
                notes="",
            )
            db.add(item)

    batch.total_amount = round(total_amount, 2)
    db.commit()
    db.refresh(batch)

    return batch_ser(batch, db, include_items=True)


# ------------------------------------------------------------------ Get single batch

@router.get("/disbursements/{branch_id}/{batch_id}/")
def get_disbursement_batch(
    branch_id: int,
    batch_id: int,
    ctx=Depends(require_perm("disbursements.view")),
):
    db: Session = ctx["db"]
    batch = db.get(DisbursementBatch, batch_id)
    if not batch or batch.branch_id != branch_id:
        raise HTTPException(status_code=404, detail="Disbursement batch not found.")

    return batch_ser(batch, db, include_items=True)


# ------------------------------------------------------------------ Update batch (draft only)

class DisbursementItemUpdateIn(BaseModel):
    id: int | None = None
    business_id: int | None = None
    percentage: float = 100.0
    is_included: bool = True
    notes: str = ""
    amount: float | None = None


class DisbursementBatchUpdateIn(BaseModel):
    title: str | None = None
    payment_date: date | None = None
    currency: str | None = None
    base_amount: float | None = None
    notes: str | None = None
    items: list[DisbursementItemUpdateIn] | None = None


@router.put("/disbursements/{branch_id}/{batch_id}/")
def update_disbursement_batch(
    branch_id: int,
    batch_id: int,
    data: DisbursementBatchUpdateIn,
    ctx=Depends(require_perm("disbursements.edit")),
):
    db: Session = ctx["db"]
    batch = db.get(DisbursementBatch, batch_id)
    if not batch or batch.branch_id != branch_id:
        raise HTTPException(status_code=404, detail="Disbursement batch not found.")

    if batch.status == "processed":
        raise HTTPException(
            status_code=400,
            detail="Cannot modify a processed disbursement batch. An administrator must reopen it first.",
        )

    if data.title is not None:
        batch.title = data.title.strip() or batch.title
    if data.payment_date is not None:
        batch.payment_date = data.payment_date
    if data.currency is not None:
        curr = data.currency.upper()
        if curr not in ("USD", "EUR", "JOD"):
            raise HTTPException(status_code=400, detail="Currency must be USD, EUR, or JOD.")
        batch.currency = curr
    if data.base_amount is not None:
        batch.base_amount = float(data.base_amount)
    if data.notes is not None:
        batch.notes = data.notes

    if data.items is not None:
        existing_items = {it.id: it for it in batch.items}
        total_amount = 0.0

        for item_data in data.items:
            if item_data.id and item_data.id in existing_items:
                item = existing_items[item_data.id]
                item.percentage = float(item_data.percentage)
                item.is_included = bool(item_data.is_included)
                item.notes = item_data.notes or ""
                calc_amt = item_data.amount if item_data.amount is not None else (batch.base_amount * (item.percentage / 100.0))
                item.amount = round(float(calc_amt), 2)
                if item.is_included:
                    total_amount += item.amount
            elif item_data.business_id:
                biz = db.get(Business, item_data.business_id)
                if biz and biz.branch_id == branch_id:
                    pct = float(item_data.percentage)
                    calc_amt = item_data.amount if item_data.amount is not None else (batch.base_amount * (pct / 100.0))
                    calc_amt = round(float(calc_amt), 2)
                    if item_data.is_included:
                        total_amount += calc_amt
                    new_item = DisbursementItem(
                        batch_id=batch.id,
                        business_id=biz.id,
                        percentage=pct,
                        amount=calc_amt,
                        is_included=item_data.is_included,
                        notes=item_data.notes or "",
                    )
                    db.add(new_item)
        batch.total_amount = round(total_amount, 2)
    else:
        # Recalculate based on updated base_amount
        total_amount = 0.0
        for item in batch.items:
            item.amount = round(float(batch.base_amount * (item.percentage / 100.0)), 2)
            if item.is_included:
                total_amount += item.amount
        batch.total_amount = round(total_amount, 2)

    batch.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(batch)

    return batch_ser(batch, db, include_items=True)


# ------------------------------------------------------------------ Confirm / Process batch

@router.post("/disbursements/{branch_id}/{batch_id}/confirm/")
def confirm_disbursement_batch(
    branch_id: int,
    batch_id: int,
    ctx=Depends(require_perm("disbursements.confirm")),
):
    db: Session = ctx["db"]
    user = ctx["user"]
    batch = db.get(DisbursementBatch, batch_id)
    if not batch or batch.branch_id != branch_id:
        raise HTTPException(status_code=404, detail="Disbursement batch not found.")

    if batch.status == "processed":
        raise HTTPException(status_code=400, detail="Batch is already processed and locked.")

    batch.status = "processed"
    batch.confirmed_by_id = user.id
    batch.confirmed_at = datetime.utcnow()
    batch.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(batch)

    return {
        "detail": "Batch confirmed and processed successfully.",
        "batch": batch_ser(batch, db, include_items=True),
    }


# ------------------------------------------------------------------ Reopen batch (Admin only)

@router.post("/disbursements/{branch_id}/{batch_id}/reopen/")
def reopen_disbursement_batch(
    branch_id: int,
    batch_id: int,
    ctx=Depends(require_branch_access()),
):
    db: Session = ctx["db"]
    roles = ctx["roles"]

    # Admin check: Must have branch_admin / organization_admin or disbursements.reopen permission
    is_admin = any(r in ("branch_admin", "organization_admin", "*") for r in roles) or has_perm(roles, "disbursements.reopen")
    if not is_admin:
        raise HTTPException(
            status_code=403,
            detail="Only administrators have permission to reopen a processed disbursement batch.",
        )

    batch = db.get(DisbursementBatch, batch_id)
    if not batch or batch.branch_id != branch_id:
        raise HTTPException(status_code=404, detail="Disbursement batch not found.")

    if batch.status == "draft":
        raise HTTPException(status_code=400, detail="Batch is already in draft status.")

    batch.status = "draft"
    batch.confirmed_by_id = None
    batch.confirmed_at = None
    batch.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(batch)

    return {
        "detail": "Batch reopened successfully.",
        "batch": batch_ser(batch, db, include_items=True),
    }


# ------------------------------------------------------------------ Delete batch (draft only)

@router.delete("/disbursements/{branch_id}/{batch_id}/")
def delete_disbursement_batch(
    branch_id: int,
    batch_id: int,
    ctx=Depends(require_perm("disbursements.edit")),
):
    db: Session = ctx["db"]
    batch = db.get(DisbursementBatch, batch_id)
    if not batch or batch.branch_id != branch_id:
        raise HTTPException(status_code=404, detail="Disbursement batch not found.")

    if batch.status == "processed":
        raise HTTPException(
            status_code=400,
            detail="Cannot delete a processed disbursement batch. Reopen it first or archive.",
        )

    db.delete(batch)
    db.commit()
    return {"detail": "Disbursement batch deleted successfully."}


# ------------------------------------------------------------------ Export batch to Excel

@router.get("/disbursements/{branch_id}/{batch_id}/export/")
def export_disbursement_batch(
    branch_id: int,
    batch_id: int,
    ctx=Depends(require_perm("disbursements.view")),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    db: Session = ctx["db"]
    batch = db.get(DisbursementBatch, batch_id)
    if not batch or batch.branch_id != branch_id:
        raise HTTPException(status_code=404, detail="Disbursement batch not found.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Disbursement Batch"

    # Header styling
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=14, bold=True)
    meta_font = Font(name="Arial", size=10, italic=True)

    # Batch Title & Meta Information
    ws.append([f"finnspark — {batch.title}"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([f"Programme / Cohort: {batch.program.name if batch.program else '—'}"])
    ws.cell(row=2, column=1).font = meta_font
    ws.append([
        f"Payment Date: {batch.payment_date} | Currency: {batch.currency} | Status: {batch.status.upper()} | Total: {batch.total_amount:,.2f} {batch.currency}"
    ])
    ws.cell(row=3, column=1).font = meta_font
    ws.append([])  # blank line

    headers = [
        "Startup / Business",
        "Founders",
        "Payment %",
        f"Disbursement Amount ({batch.currency})",
        "Included in Batch",
        "Remarks / Notes",
    ]
    ws.append(headers)
    header_row_idx = 5
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if col_idx in (3, 4, 5) else "left", vertical="center")

    for item in batch.items:
        founders_str = ", ".join(
            f"{f.first_name} {f.last_name}".strip() for f in (item.business.founders if item.business else [])
        )
        row_data = [
            item.business.name if item.business else "—",
            founders_str,
            f"{item.percentage:.1f}%",
            item.amount,
            "Yes" if item.is_included else "No",
            item.notes or "",
        ]
        ws.append(row_data)

    # Summary Row
    summary_row = [
        "TOTAL COHORT DISBURSEMENT",
        "",
        "",
        batch.total_amount,
        f"{sum(1 for i in batch.items if i.is_included)} / {len(batch.items)} startups",
        "",
    ]
    ws.append(summary_row)
    last_row = ws.max_row
    total_font = Font(name="Arial", size=11, bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.font = total_font

    # Column widths auto-adjust
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"disbursement_{batch.id}_{batch.payment_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
